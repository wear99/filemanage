
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill
from openpyxl.comments import Comment
from parts.search import Partfind_dict
from parts.read_excel import read_design_BOM

import re

task_status = {}
def check_code(code, draw, name):
    '对物料编码进行确认：如果输入了编码，则根据末尾和图纸查找是不是最新版；如果没有输入编码，则先根据图号查找有无唯一编码，若无则根据名称查找有无唯一编码。如果查找到的是同一个编码不同版本，则返回最新编码'
    def rst_check(rst):
        new_code = ''
        codes = []
        for i in rst:
            if re.match(r'\d{2}R\d{7}', i['code']):  # 挑选出有真正编码的，去掉设计物料
                codes.append(i['code'])

        if codes:
            codes.sort()
            codes.reverse()
            if len(codes) == 1:
                new_code = codes[0]
            elif codes.count(codes[0][:-1]) == len(codes):  # 当查询的编码除最后一位都相同时
                new_code = codes[0]

        return new_code

    new_code = ''
    rst = {}
    rst_ = {}
    if code==draw+name:
        code=''
    rst['old'] = code
    
    if draw != '' and 'GB' not in str(draw): #先用图号查找
        rst_ = Partfind_dict(type='draw', search=(draw,),opt='and')
        if rst_:
            new_code = rst_check(rst_)

    if not new_code:  #如果没有结果，再用名称查找唯一结果
        rst_ = Partfind_dict(type='name', search=(name,), opt='and')
        if rst_:
            new_code = rst_check(rst_)

    if not new_code or new_code == code:  # 当没找到新的编码时或者查找结果和原来相同，返回旧编码
        pass

    elif not code or code == ' ':  # 当原编码为空时，直接写入
        rst['new'] = new_code
        rst['style'] = 'green'

    else:  # 当原编码存在时，进行对比
        if code[:-1] == new_code[:-1]:  # 当找到的新旧编码除最后一位外相同时
            if code[-1] < new_code[-1]:
                rst['new'] = new_code
                rst['style'] = 'blue'
                rst['remark'] = '原编码为老版本：'+str(code)
            else:
                pass
        else:
            rst['new'] = new_code
            rst['style'] = 'red'
            rst['remark'] = '由图号查询的编码不同，原内容:\n'+str(code)

    return rst


def update_excel(fpath, bom_dict):
    '根据输入字典 { 行#列：{内容,备注,样式},  insert_col: ,}'
    fill_style = {'blue': PatternFill('solid', fgColor='87CEEB'),
                    'red': PatternFill('solid', fgColor='FF4500'),
                    'green': PatternFill('solid', fgColor='00FF7F'),
                    'oringe': PatternFill('solid', fgColor='FFA500'),
                    'gray': PatternFill('solid', fgColor='A9A9A9'),
                    }
    rst = {}
    try:
        wbook = load_workbook(fpath, read_only=False)
        wsheet = wbook.active
    except Exception as ex:
        rst['error'] = '文件读取失败：'+str(ex)
        return rst

    if 'insert_col' in bom_dict:
        # 在数量列后面插入一列,inset会插入在前面,cols时列从1开始，而rows数组从0开始
        wsheet.insert_cols(bom_dict['insert_col'])
    if 'update' in bom_dict:
        for key, val in bom_dict['update'].items():
            row = key[0]
            col = key[1]
            if 'new' in val:
                wsheet.cell(row, col).value = val['new']
            else:
                wsheet.cell(row, col).value = val['old']

            if 'remark' in val:
                wsheet.cell(row, col).comment = Comment(
                    val['remark'], 'check')
            if 'style' in val:
                wsheet.cell(row, col).fill = fill_style[val['style']]

    try:
        wbook.save(fpath)
    except Exception as ex:
        rst['error'] = 'EXCEL文件写入失败：' + str(ex)
    finally:
        wbook.close()

    return rst


def check_excel_code(fpath,task_id):  # 查找excel表编码
    def check_bom(bom, col):
        rst = {'update':{}}
        checked = {}
        rst['insert_col'] = col
        for item in bom:
            if (item['code'], item['draw'], item['name']) in checked:  # 对于已查找过物料，直接使用之前查询结果
                code = checked[item['code'], item['draw'], item['name']]
            else:
                code = check_code(item['code'], item['draw'], item['name'])
                checked[item['code'], item['draw'], item['name']] = code

            rst['update'][(item['row'], col)] = code
        return rst

    def check_QTY(bom, col):
        rst = {}
        lv_num = {0: 1}
        for item in bom:
            lv = item[0]
            num = item[2]
            for n in range(lv + 1, 7):
                lv_num[n] = 0
            lv_num[lv] = num * lv_num[lv - 1]
            rst[str(item[-1])+'#'+str(col)
                ] = {'new': lv_num[lv], 'old': num}
        return rst

    task_status[task_id].append({'task': '读取表格内容', 'rst': ''})
    # 读取excel内容
    rst = read_design_BOM(fpath, type='CHECK')
    if 'error' in rst:
        return rst

    #对文件进行检查
    task_status[task_id].append({'task': '检查编码', 'rst': ''})
    col = rst['col']['code']+2
    check_rst = check_bom(rst['bom'], col)    
    if 'error' in check_rst:
        return check_rst

    task_status[task_id][-1]['rst']='完成'
    #对文件进行更新
    task_status[task_id].append({'task': '更新表格', 'rst': ''})
    update_rst = update_excel(fpath, check_rst)
    if 'error' in update_rst:
        return update_rst
    task_status[task_id][-1]['rst'] = '完成'
    # 检查编码    
    return update_rst

