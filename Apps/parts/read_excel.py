from openpyxl import load_workbook
import re
from datetime import datetime
# 读取上传的excel文件,格式化数据，找到对应的文件,存入数据库
rule = {'code': r'\d{2}R\d{7}',
        'batch': r'C|E\d{2}-|D',
        'asmb': r'24|8R',
        'metal': r'0[1234]R',
        'root': r'(:?C|E|N|YF|EXP|CS)\d{2}(:?-|D|\d{2})',
        'design': r'N\d{4}',
        'custom': r'YF\d{7}',
        'exper': r'(EXP|CS)\d{7}',
        'temp': r'TEMP\d{2}'
        }


def read_design_BOM(fpath, type='BATCH'):
    '统一读取各种excel文件,根据不同类型生成不同列；读取时进行序号、编码、数量格式检查；针对check和计算数量模式会写入到原excel'
    # [0层次,1编码,2图号,3名称,4数量,5材料,6重量,7备注,8材料成本,9人工成本,10管理成本]
    def get_col(wsheet):
        col_ = {}
        lable = {
            '级别': 'lv',
            '层次': 'lv',
            '序号': 'lv',
            '序列号': 'sn',
            '子件编码': 'code',
            '新编码': 'code',
            '编码': 'code',
            '代号': 'draw',
            '图号': 'draw',
            '子件名称': 'name',
            '描述': 'name',
            '名称': 'name',
            '用量': 'quantity',
            '基本用量': 'quantity',
            '使用数量': 'quantity',
            '数量': 'quantity',
            '子件规格': 'material',
            '材料': 'material',
            '单重': 'weight',
            '备注': 'remark',
            '材料成本': 'material_cost',
            '人工成本': 'labor_cost',
            '费用成本': 'managed_cost',
            '日期': 'time',
            '时间': 'time',
            '零部件图号': 'draw',
            '零部件名称': 'name',
            '更改前编码': 'before_code',
            '更改后编码': 'after_code',
            '更改前说明': 'before_des',
            '更改后说明': 'after_des',
            '更改类别': 'change_type',
            '更改方式': 'change_draw',
            '库存': 'stock',
            '在途': 'on_order',
            '已制品处理建议': 'suggestion',
            '涉及机型': 'product',
            '分工': 'division',
            '发放部门': 'output',
        }

        lable_t = {}
        for key, item in lable.items():
            if item in title:
                lable_t[key] = item

        for row in wsheet.values:  # 只检查第一行
            col_['used'] = []
            for c, value in enumerate(row):
                for key in lable_t:
                    if isinstance(value, str) and key == value.replace(' ', ''):
                        col_['used'].append(c)
                        if lable_t[key] == 'lv':
                            if 'lv' not in col_:
                                col_['lv'] = [c, ]
                            else:
                                col_['lv'].append(c)
                        else:
                            if lable_t[key] in col_:
                                return '%s 的属性列重复' % key
                            else:
                                col_[lable_t[key]] = c

            name = {val: key for key, val in lable_t.items()}
            for item in need:
                if item not in col_:
                    return '找不到属性列: '+ name[item]
            return col_

    def read_item(wsheet):
        def fmt_str(x):
            if x:
                x = str(x)
                x = x.replace(' ', '')
                if x:
                    return x.upper()
                else:
                    return ''
            else:
                return ''

        def fmt_num(x):
            if not x or x == ' ':
                return 0
            elif isinstance(x, int):
                return x
            else:
                try:
                    return round(float(x), 2)
                except:
                    return x

        def code_rule(x):
            if re.match(rule['code'], x):
                return True
            elif re.match(rule['root'], x):
                return True
            elif re.match(rule['temp'], x):
                return True
            else:
                return False

        def fmt_time(x):
            if isinstance(x,datetime):
                return x
            else:
                return ""

        def get_lv(lvs):
            if len(lvs) == 1:  # 针对层次只有1列：一种是'+++'，另一种是数字加.来区分
                if not lvs[0]:
                    return ' 缺少层次'

                lv1 = str(lvs[0])
                if '+' in lv1:
                    lv = len(lv1)
                else:
                    lv = lv1.count('.')+2
                return lv
            else:   # 层次由多列，EBOM格式，数字所在的列代表层次高低
                lv_1 = []
                for col, n in enumerate(lvs):
                    if isinstance(n, int):
                        lv = col+1
                        lv_1.append(lv)
                if len(lv_1) == 1:
                    return lv
                elif len(lv_1) > 1:
                    return ' 行层次重复'
                else:
                    return ' 行缺少层次'

        row_num = 1
        lv_2 = 2
        root_row=0
        for row in wsheet.iter_rows(values_only=True,min_row=2):
            row_num += 1
            item = {}
            item['row'] = row_num
            if 'lv' in title and row_num == 2:
                if row[0] and row[0].upper() == 'ROOT':
                    pass
                else:
                    item_error.append('第 ' + str(row_num) + '没有ROOT')
                    break

            for key in title:
                if key in col:
                    if key == 'lv':
                        if row[0] and row[0].upper() == 'ROOT':
                            item['lv'] =1
                            root_row = row_num
                        else:
                            item['lv'] = get_lv([row[x] for x in col['lv']])

                    elif key in ('quantity', 'weight', 'material_cost', 'labor_cost', 'managed_cost'):
                        item[key] = fmt_num(row[col[key]])

                    elif key in ('time',):
                        item[key] = fmt_time(row[col[key]])

                    else:
                        item[key] = fmt_str(row[col[key]])

                else:
                    item[key] = ''

            if 'name' in item and item['name'] == '':  # 跳过空行
                continue

            if 'lv' in item:
                if isinstance(item['lv'], int):
                    if row_num == root_row+1:   # Root的第二行的层次必须是2, 确定一个层次的调整系数
                        lv_2 = item['lv']
                    item['lv'] = item['lv']+2-lv_2

                    # 检查层次是否脱节
                    if len(excel_bom) > 1 and item['lv'] > excel_bom[-1]['lv'] + 1:
                        item_error.append('第 ' + str(row_num) + ' 行层次和上层脱节')
                else:
                    item_error.append('第 ' + str(row_num) + item['lv'])
                    item['lv']=9

            if 'quantity' in item:
                if not isinstance(item['quantity'], (int, float)):
                    item_error.append('第 ' + str(row_num) + '行没有数量或格式不对')

            if 'material_cost' in item:
                try:
                    item['cost'] = float(item['material_cost']) + float(item['labor_cost'])+float(item['managed_cost'])
                except:
                    item_error.append('第 ' + str(row_num) + '行成本不是数字格式')

            if 'code' in item:  # 对编码格式进行检查
                if not item['code']:   # 当编码为空时
                    if type in ('ERPBOM'):
                        item_error.append('第 ' + str(row_num) + '行缺少编码')
                    else:
                        item['code'] = item['draw']+item['name']

                elif not code_rule(item['code']):  # 当编码格式不正确时
                    if type in ('TEMP', 'REMOVE', 'COST',):  # 不处理
                        pass
                    else:
                        item_error.append('第 ' + str(row_num) + '行编码格式不对')

            if not item_error:
                # 物料字段以字典形式
                excel_bom.append(item)

    def add_sn(read_bom):
        sn = {1: 0}
        new = []
        for item in read_bom:
            if not isinstance(item['lv'], int):
                return
            lv = item['lv']
            # 添加SN
            sn[item['lv']] += 1
            s = ''
            for n in range(1, 8):
                if n <= item['lv']:
                    if s:
                        s = s+'.'+str(sn[n])
                    else:
                        s = str(sn[n])
                else:
                    sn[n] = 0
            item['sn'] = s
            new.append(item)

        return new

    def add_parent(read_bom,field):
        pids={0:'ROOT'}
        new=[]
        for item in read_bom:
            if not isinstance(item['lv'], int):
                return
            if field not in item:
                return

            lv=item['lv']
            # 添加父项字段
            pids[lv] = item[field]
            item['parent'] = pids[lv-1]

            new.append(item)

        return new

    def add_total(read_bom):
        lv={0:1}

        for item in read_bom:
            if not isinstance(item['lv'], int):
                return

            lv[item['lv']] = lv[item['lv']-1]*item['quantity']
            item['total'] = lv[item['lv']]

            for n in range(item['lv'] + 1, 8):
                lv[n] = 1
        return read_bom


    excel_bom = []    
    rst = {}
    item_error = []

    if type in ('BATCH', 'DESIGN', 'EXPER', 'CUSTOM', 'ARCHIVE','ERPBOM'):
        title = ('lv', 'code', 'draw', 'name',
                 'quantity', 'material', 'weight', 'remark', 'division', 'output')
        need = ('lv', 'code', 'name','quantity')
    elif type == 'CODE':
        title = ('code', 'draw', 'name', 'material',
                 'weight', 'remark', 'time')
        need = ('code', 'name')
    elif type == 'COST':
        title = ('code', 'name', 'material_cost', 'labor_cost', 'managed_cost','cost')
        need = ('code', 'name', 'material_cost', 'labor_cost', 'managed_cost')
    elif type == 'CHANGE':
        title = ('lv', 'draw', 'name', 'before_code', 'after_code', 'before_des', 'after_des',
                 'change_type', 'change_draw', 'stock', 'on_order', 'suggestion', 'product',)
        need = ('sn', 'name', 'before_code', 'after_code', 'change_type')
    elif type == 'CHECK':
        title = ('code', 'draw', 'name')
        need = ('code', 'draw', 'name')
    elif type == 'QTY':
        title = ('lv', 'name', 'quantity')
        need = ('lv', 'name', 'quantity')
    elif type in ('TEMP', 'REMOVE'):
        title = ('lv', 'code', 'draw', 'name', 'quantity')
        need = ('lv', 'code', 'name', 'quantity')
    elif type == 'APPLICATION':
        title = ('lv','code', 'draw', 'name')
        need = ('code', 'draw','name')
    else:
        return {}

    try:
        wbook = load_workbook(fpath, read_only=True,data_only=True)
    except Exception as ex:
        rst['error'] = '文件读取失败：'+str(ex)
        return rst

    if type == 'CODE':
        names=wbook.sheetnames
    else:
        names = [wbook.active.title]

    for sname in names:  #读取多个工作表
        wsheet = wbook[sname]
        col=get_col(wsheet)
        if isinstance(col, dict):
            rst['col']=col
            read_item(wsheet)
        else:
            if len(names)>1:
                skip_sheet += wsheet.title +' : '+col+ ' ;'
            else:
                rst['error'] = '缺少标题行：' + skip_sheet

    wbook.close()
    if skip_sheet:
        rst['skip'] = '跳过的工作表：'+skip_sheet        

    if item_error:
        item_error = [[1, x] for x in item_error]
        rst['error'] = item_error

    if 'error' not in rst:
        return rst

    if excel_bom:
        if 'lv' in title:
            excel_bom = add_sn(excel_bom)
            if 'code' in need:
                excel_bom = add_parent(excel_bom,'code')
            elif 'before_code' in need:
                excel_bom = add_parent(excel_bom, 'sn')

            if 'quantity' in need:
                excel_bom = add_total(excel_bom)

        rst['bom'] = excel_bom        

    return rst  # {字典形式}


